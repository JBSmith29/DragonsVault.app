# services/csv_importer.py
import csv
import io
import logging
import os
import uuid
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Any, Tuple, Optional, Iterable
from sqlalchemy.exc import IntegrityError
from sqlalchemy import or_, func
from flask import current_app, has_request_context
from extensions import db, cache
from models import Card, Folder
from pathlib import Path
from services.scryfall_cache import cache_exists, load_cache, find_by_set_cn, metadata_from_print
from services.live_updates import emit_import_event

# --- NEW: lightweight struct for preview data ---
@dataclass
class PreviewData:
    headers: List[str]
    rows: List[Dict[str, Any]]
# ------------------------------------------------

# --- NEW: lazy-load helper for the Scryfall cache ---
_CACHE_READY: Optional[bool] = None
IMPORT_ENABLE_SCRYFALL_LOOKUPS = os.getenv("IMPORT_ENABLE_SCRYFALL_LOOKUPS", "1").lower() in {"1", "true", "yes", "on"}
def _ensure_cache_loaded() -> bool:
    """Load Scryfall bulk cache once per process, if present."""
    global _CACHE_READY
    if _CACHE_READY is None:
        if not IMPORT_ENABLE_SCRYFALL_LOOKUPS:
            logger.warning("Skipping Scryfall cache load for CSV import (IMPORT_ENABLE_SCRYFALL_LOOKUPS=0).")
            _CACHE_READY = False
            return False
        _CACHE_READY = cache_exists() and load_cache()
    return bool(_CACHE_READY)
# ----------------------------------------------------

_METADATA_FIELDS = ("type_line", "rarity", "color_identity_mask")
IMPORT_CHUNK_SIZE = int(os.getenv("IMPORT_BATCH_SIZE", 500))

logger = logging.getLogger(__name__)


class HeaderValidationError(ValueError):
    """Raised when required headers are missing from the import file."""

    def __init__(self, details: List[str]):
        super().__init__("Missing required column(s): " + "; ".join(details))
        self.details = details


def _apply_card_metadata(card: Card, metadata: Dict[str, Any], *, only_if_empty: bool = False) -> bool:
    """Apply derived metadata to a Card, returning True if any field changed."""
    changed = False
    for field in _METADATA_FIELDS:
        value = metadata.get(field)
        if value is None:
            continue
        current = getattr(card, field, None)
        if only_if_empty and current not in (None, "", 0):
            continue
        if current != value:
            setattr(card, field, value)
            changed = True
    return changed

# Ordered header variants (first match wins).
EXPECTED = {
    "folder": [
        "folder name", "folder_name", "folder", "binder", "binder name", "album",
    ],
    "folder_category": [
        "folder category", "folder type", "binder type", "collection type",
    ],
    "name": [
        "card name", "card_name", "name", "card",
    ],
    "qty": [
        "quantity", "qty", "trade quantity", "count", "copies",
    ],
    "set_code": [
        "set code", "set_code", "set", "expansion", "setcode", "edition",
    ],
    "collector_number": [
        "collector number", "collector_number", "collector #",
        "card number", "card_number", "card #",
        "cn", "number", "#",
    ],
    "lang": [
        "language", "lang",
    ],
    "is_foil": [
        "printing", "foil", "is foil", "foil?", "is_foil",
    ],
}

_LANG_MAP = {
    "english": "en",
    "japanese": "ja",
    "spanish": "es",
    "german": "de",
    "french": "fr",
    "italian": "it",
    "portuguese": "pt",
    "korean": "ko",
    "russian": "ru",
    "simplified chinese": "zhs",
    "chinese simplified": "zhs",
    "traditional chinese": "zht",
    "chinese traditional": "zht",
    "phyrexian": "ph",
}

def _norm_lang(v: Optional[str]) -> str:
    if not v:
        return "en"
    s = str(v).strip().lower()
    return _LANG_MAP.get(s, s[:5]) or "en"


def _is_moxfield_headers(headers: Iterable[str]) -> bool:
    """Heuristically detect a Moxfield export."""
    lowers = {h.strip().lower() for h in headers if isinstance(h, str)}
    required = {"count", "name", "edition"}
    return required.issubset(lowers) and ("purchase price" in lowers or "alter" in lowers)

_POSITIVE_FOIL_VALUES = {
    "1",
    "true",
    "t",
    "y",
    "yes",
    "foil",
    "foils",
    "foilonly",
    "etched",
    "etch",
    "gilded",
    "gild",
    "glossy",
    "textured",
    "neon",
    "neonink",
    "halo",
    "halofoil",
    "surge",
    "surgefoil",
    "galaxy",
    "cosmic",
    "rainbow",
    "shiny",
    "sparkle",
    "sparkly",
    "prismatic",
    "oil",
    "oilfoil",
    "stepandcompleat",
    "raised",
    "embossed",
}

_NEGATIVE_FOIL_VALUES = {
    "0",
    "false",
    "f",
    "n",
    "no",
    "nonfoil",
    "nonfoils",
    "non-foil",
    "nf",
    "normal",
    "regular",
}

_POSITIVE_FOIL_SUBSTRINGS = (
    "foil",
    "etched",
    "gild",
    "gloss",
    "textur",
    "neon",
    "halo",
    "surge",
    "galaxy",
    "cosmic",
    "rainbow",
    "spark",
    "prism",
    "oil",
    "step-and-compleat",
    "stepandcompleat",
    "raised",
    "emboss",
)

_NEGATIVE_FOIL_SUBSTRINGS = (
    "nonfoil",
    "non-foil",
    "non foil",
    "non/foil",
    "no foil",
)


def _to_bool(v: Any) -> bool:
    if v is None:
        return False
    s = str(v).strip().lower()
    if not s:
        return False

    normalized = s.replace(" ", "").replace("-", "").replace("_", "")
    if normalized in _NEGATIVE_FOIL_VALUES:
        return False
    if normalized in _POSITIVE_FOIL_VALUES:
        return True

    for token in _POSITIVE_FOIL_SUBSTRINGS:
        if token in s:
            return True
    for token in _NEGATIVE_FOIL_SUBSTRINGS:
        if token in s:
            return False
    return False

def _to_int(v: Any, default=1) -> int:
    try:
        return int(str(v).strip())
    except Exception:
        return default

def _norm_set_code(v: str) -> str:
    return (v or "").strip().lower()

def _norm_cn(v: Any) -> str:
    return str(v).strip()

_FOLDER_CATEGORY_KEYWORDS = {
    "deck": Folder.CATEGORY_DECK,
    "decks": Folder.CATEGORY_DECK,
    "decklist": Folder.CATEGORY_DECK,
    "commander": Folder.CATEGORY_DECK,
    "binder": Folder.CATEGORY_COLLECTION,
    "collection": Folder.CATEGORY_COLLECTION,
    "trade": Folder.CATEGORY_COLLECTION,
    "binder/trade": Folder.CATEGORY_COLLECTION,
}

def _norm_folder_category(raw: Optional[str]) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if not s:
        return None
    if s in _FOLDER_CATEGORY_KEYWORDS:
        return _FOLDER_CATEGORY_KEYWORDS[s]
    if "deck" in s:
        return Folder.CATEGORY_DECK
    if any(tok in s for tok in ("binder", "collection", "trade", "inventory")):
        return Folder.CATEGORY_COLLECTION
    return None


def _folder_query_for_owner(name: str, owner_user_id: Optional[int]):
    """Return a query restricted to the owning user's folder namespace (case-insensitive)."""
    query = Folder.query.filter(func.lower(Folder.name) == func.lower(name))
    if owner_user_id is not None:
        query = query.filter(Folder.owner_user_id == owner_user_id)
    return query


def _ensure_folder_for_owner(
    folder_name: str,
    folder_category: Optional[str],
    owner_user_id: Optional[int],
    owner_name: Optional[str],
) -> Folder | None:
    """
    Return an existing folder for this owner/name or create one safely.
    Handles duplicate names by reusing or generating a unique variant.
    """
    folder = _folder_query_for_owner(folder_name, owner_user_id).first()
    if folder:
        if folder_category and folder.category != folder_category:
            folder.category = folder_category
        if owner_user_id and not folder.owner_user_id:
            folder.owner_user_id = owner_user_id
        if owner_name and (not folder.owner or folder.owner_user_id == owner_user_id):
            folder.owner = owner_name
        return folder

    folder = Folder(name=folder_name)
    if folder_category:
        folder.category = folder_category
    if owner_user_id:
        folder.owner_user_id = owner_user_id
    if owner_name:
        folder.owner = owner_name
    db.session.add(folder)
    try:
        db.session.flush()
        return folder
    except IntegrityError:
        db.session.rollback()
        existing = _folder_query_for_owner(folder_name, owner_user_id).first()
        if existing:
            return existing
        suffix = 2
        base = folder_name
        while True:
            candidate = f"{base} ({suffix})"
            conflict = _folder_query_for_owner(candidate, owner_user_id).first()
            if not conflict:
                alt = Folder(name=candidate)
                if folder_category:
                    alt.category = folder_category
                if owner_user_id:
                    alt.owner_user_id = owner_user_id
                if owner_name:
                    alt.owner = owner_name
                db.session.add(alt)
                db.session.flush()
                return alt
            suffix += 1
    except Exception:
        db.session.rollback()
        raise

def _read_text(filepath: str) -> str:
    with open(filepath, "r", encoding="utf-8-sig", errors="replace") as f:
        return f.read()

def _strip_sep_preamble(text: str) -> Tuple[str, Optional[str]]:
    """
    Handles Excel 'sep=,' preamble, quoted or unquoted.
    Returns (content_without_preamble, delimiter_if_declared)
    """
    if not text:
        return text, None
    first_newline = text.find("\n")
    head = text if first_newline == -1 else text[:first_newline]
    head_stripped = head.strip().lower().strip("'\"")
    if head_stripped.startswith("sep=") and len(head_stripped) >= 5:
        delim = head_stripped[4:5]
        content = text[first_newline + 1 :] if first_newline != -1 else ""
        return content, delim
    return text, None

def _make_reader(filepath: str) -> Tuple[csv.DictReader, str]:
    raw = _read_text(filepath)
    content, declared = _strip_sep_preamble(raw)

    if declared:
        delimiter = declared
    else:
        sample = content[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
            delimiter = dialect.delimiter
        except Exception:
            delimiter = ","

    sio = io.StringIO(content)
    reader = csv.DictReader(sio, delimiter=delimiter)
    return reader, delimiter

def _normalize_headers(headers):
    if not headers:
        raise HeaderValidationError([
            "No headers found. Include columns such as 'Name', 'Set Code', 'Collector Number'."
        ])
    lower_to_original = {h.strip().lower(): h for h in headers if isinstance(h, str)}
    mapping = {}
    for field, variants in EXPECTED.items():
        for v in variants:
            if v in lower_to_original:
                mapping[field] = lower_to_original[v]
                break
    required_fields = {
        "name": "Card name",
        "set_code": "Set code",
        "collector_number": "Collector number",
    }
    missing = []
    for req, label in required_fields.items():
        if req not in mapping:
            variants = ", ".join(EXPECTED.get(req, []))
            detail = f"{label} (accepted: {variants})" if variants else label
            missing.append(detail)
    if missing:
        raise HeaderValidationError(missing)
    return mapping


def validate_import_file(filepath: str) -> None:
    """Ensure the provided file has the required headers."""
    _, headers, _ = _open_table(filepath)
    _normalize_headers(headers or [])

@dataclass
class ImportStats:
    added: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    job_id: Optional[str] = None
    skipped_details: list = field(default_factory=list)


SKIP_DETAIL_LIMIT = 50

def process_csv(
    filepath: str,
    default_folder: str = "Unsorted",
    dry_run: bool = False,
    quantity_mode: str = "new_only",  # default: add only new rows; legacy modes remain for backwards compatibility
    *,
    job_id: Optional[str] = None,
    owner_user_id: Optional[int] = None,
    owner_username: Optional[str] = None,
) -> Tuple[ImportStats, Dict[str, int]]:
    """
    Reads a CSV or Excel file and upserts Cards. Returns (stats, by_folder_counts).
    - quantity_mode="new_only": skip existing matches, add only new prints (default).
    - quantity_mode="delta": add CSV qty onto existing (legacy; not exposed in UI).
    - quantity_mode="purge": same as delta, but the caller cleared cards first.
    - quantity_mode="absolute": legacy replace mode (set quantity to CSV value).
    """
    stats = ImportStats()
    stats.job_id = job_id or uuid.uuid4().hex
    job_id = stats.job_id
    source_name = Path(filepath).name
    emit_import_event(
        "started",
        job_id=job_id,
        file=source_name,
        dry_run=dry_run,
        default_folder=default_folder,
        quantity_mode=quantity_mode,
    )
    per_folder: Dict[str, int] = {}

    rows_iter, headers, _delimiter = _open_table(filepath)  # CHANGED: unified loader
    is_moxfield = _is_moxfield_headers(headers or [])
    default_folder_local = "Collection" if is_moxfield else default_folder
    processed = 0
    mapping = _normalize_headers(headers or [])

    owner_name = (owner_username or "").strip() or None

    def _commit_batch():
        try:
            db.session.commit()
            if not has_request_context():
                # Releasing the identity map keeps worker memory flat without detaching web request users.
                db.session.expunge_all()
        except Exception:
            db.session.rollback()
            raise

    for row in rows_iter:
        try:
            folder_name = (row.get(mapping.get("folder")) or default_folder_local).strip()
            if is_moxfield:
                folder_category = Folder.CATEGORY_COLLECTION
            else:
                folder_category = _norm_folder_category(
                    row.get(mapping["folder_category"]) if "folder_category" in mapping else None
                )
            name = (row.get(mapping["name"]) or "").strip()
            if not name:
                stats.skipped += 1
                if len(stats.skipped_details) < SKIP_DETAIL_LIMIT:
                    stats.skipped_details.append(
                        {"reason": "Missing name", "row": row}
                    )
                continue

            qty = _to_int(row.get(mapping.get("qty"), 1), default=1)
            set_code = _norm_set_code(row.get(mapping["set_code"]))
            collector_number = _norm_cn(row.get(mapping["collector_number"]))
            lang_raw = row.get(mapping.get("lang"), "en")
            lang = _norm_lang(lang_raw)
            is_foil = _to_bool(row.get(mapping.get("is_foil")))

            # ensure folder
            folder = None
            if not dry_run:
                folder = _ensure_folder_for_owner(folder_name, folder_category, owner_user_id, owner_name)
            else:
                folder = _folder_query_for_owner(folder_name, owner_user_id).first()

            key = dict(
                name=name,
                folder_id=folder.id if folder else None,
                set_code=set_code,
                collector_number=collector_number,
                lang=lang,
                is_foil=is_foil,
            )

            found = None
            metadata: Dict[str, Any] = {}

            # Enrich with oracle_id/metadata from Scryfall cache (if available)
            if not dry_run and _ensure_cache_loaded():
                found = find_by_set_cn(set_code, collector_number, name)
                metadata = metadata_from_print(found)
                if found:
                    key["oracle_id"] = found.get("oracle_id")
                else:
                    current_app.logger.warning(
                        "Import: no Scryfall match for %s [%s %s] (lang=%s, foil=%s)",
                        name,
                        set_code,
                        collector_number,
                        lang,
                        is_foil,
                    )

            # DRY-RUN path computes what would happen
            if dry_run:
                existed = False
                if folder:
                    existed = Card.query.filter_by(**key).first() is not None
                if not existed:
                    stats.added += 1
                else:
                    if quantity_mode == "absolute":
                        # We need the current quantity to decide updated vs skipped
                        existing = Card.query.filter_by(**key).first()
                        if existing and (existing.quantity or 0) != qty:
                            stats.updated += 1
                        else:
                            stats.skipped += 1
                            if len(stats.skipped_details) < SKIP_DETAIL_LIMIT:
                                stats.skipped_details.append(
                                    {
                                        "reason": "Unchanged quantity (absolute mode)",
                                        "name": name,
                                        "set_code": set_code,
                                        "collector_number": collector_number,
                                        "folder": folder_name,
                                    }
                                )
                    elif quantity_mode == "new_only":
                        stats.skipped += 1
                        if len(stats.skipped_details) < SKIP_DETAIL_LIMIT:
                            stats.skipped_details.append(
                                {
                                    "reason": "Existing card (new only mode)",
                                    "name": name,
                                    "set_code": set_code,
                                    "collector_number": collector_number,
                                    "folder": folder_name,
                                }
                            )
                    else:  # delta always updates
                        stats.updated += 1
                per_folder[folder_name] = per_folder.get(folder_name, 0) + qty
                processed += 1
                if processed % 25 == 0:
                    emit_import_event(
                        "progress",
                        job_id=job_id,
                        file=source_name,
                        processed=processed,
                        stats=asdict(stats),
                    )
                continue

            # Upsert (with row-level minimality)
            existing = Card.query.filter_by(**key).with_for_update(of=Card).first()
            if existing:
                if quantity_mode == "new_only":
                    stats.skipped += 1
                    if len(stats.skipped_details) < SKIP_DETAIL_LIMIT:
                        stats.skipped_details.append(
                            {
                                "reason": "Existing card (new only mode)",
                                "name": name,
                                "set_code": set_code,
                                "collector_number": collector_number,
                                "folder": folder_name,
                            }
                        )
                    continue
                changed = False

                # Fill oracle_id if missing
                if (existing.oracle_id in (None, "")) and _ensure_cache_loaded():
                    found = find_by_set_cn(set_code, collector_number, name)
                    if found:
                        existing.oracle_id = found.get("oracle_id")
                        metadata = metadata_from_print(found)
                        changed = True

                if metadata:
                    changed = _apply_card_metadata(existing, metadata) or changed

                if quantity_mode == "absolute":
                    if (existing.quantity or 0) != qty:
                        existing.quantity = qty
                        changed = True
                else:  # "delta"
                    existing.quantity = (existing.quantity or 0) + qty
                    changed = True

                if changed:
                    stats.updated += 1
                else:
                    stats.skipped += 1
                    if len(stats.skipped_details) < SKIP_DETAIL_LIMIT:
                        stats.skipped_details.append(
                            {
                                "reason": "Unchanged quantity (absolute mode)",
                                "name": name,
                                "set_code": set_code,
                                "collector_number": collector_number,
                                "folder": folder_name,
                            }
                        )
            else:
                # New row
                initial_qty = qty if quantity_mode == "absolute" else max(qty, 0)
                card_kwargs = {
                    **key,
                    "quantity": initial_qty,
                    "type_line": metadata.get("type_line"),
                    "rarity": metadata.get("rarity"),
                    "color_identity_mask": metadata.get("color_identity_mask"),
                }
                card = Card(**card_kwargs)
                db.session.add(card)
                try:
                    db.session.flush()
                    stats.added += 1
                except IntegrityError:
                    # Rare race—retry as update
                    db.session.rollback()
                    existing = Card.query.filter_by(**key).first()
                    if existing:
                        if (existing.oracle_id in (None, "")) and _ensure_cache_loaded():
                            found = find_by_set_cn(set_code, collector_number, name)
                            if found:
                                existing.oracle_id = found.get("oracle_id")
                                metadata = metadata_from_print(found)
                                _apply_card_metadata(existing, metadata)
                        if quantity_mode == "absolute":
                            if (existing.quantity or 0) != qty:
                                existing.quantity = qty
                                stats.updated += 1
                            else:
                                stats.skipped += 1
                        else:
                            existing.quantity = (existing.quantity or 0) + qty
                            stats.updated += 1
                    else:
                        stats.errors += 1
                        current_app.logger.exception("IntegrityError without recover", exc_info=True)

            per_folder[folder_name] = per_folder.get(folder_name, 0) + qty
            processed += 1
            if processed % 25 == 0:
                emit_import_event(
                    "progress",
                    job_id=job_id,
                    file=source_name,
                        processed=processed,
                        stats=asdict(stats),
                    )
        except Exception:
            stats.errors += 1
            current_app.logger.exception("Error processing row", exc_info=True)
        # Commit in batches to reduce memory usage on large imports
        if not dry_run and IMPORT_CHUNK_SIZE > 0 and processed and processed % IMPORT_CHUNK_SIZE == 0:
            _commit_batch()

    if processed and processed % 25:
        emit_import_event(
            "progress",
            job_id=job_id,
            file=source_name,
            processed=processed,
            stats=asdict(stats),
        )

    if not dry_run:
        _commit_batch()
        try:
            cache.clear()
        except Exception:
            pass
    emit_import_event(
        "completed",
        job_id=job_id,
        file=source_name,
        processed=processed,
        dry_run=dry_run,
        stats=asdict(stats),
        per_folder=per_folder,
    )
    _log_import_summary(
        stats,
        per_folder,
        filepath,
        dry_run=dry_run,
        quantity_mode=quantity_mode,
    )
    return stats, per_folder

def _is_excel(filepath: str) -> bool:
    return Path(filepath).suffix.lower() in {".xlsx", ".xlsm"}

def _iter_excel_rows(filepath: str) -> tuple[Iterable[Dict[str, Any]], list[str]]:
    """
    Stream rows from the first worksheet of an .xlsx/.xlsm file.
    Returns (row_iter, headers). All values are coerced to str ('' for None).
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    # Find header row (first non-empty row)
    headers = None
    for row in ws.iter_rows(values_only=True):
        vals = ["" if v is None else str(v).strip() for v in row]
        if any(vals):
            headers = vals
            break
    if not headers:
        raise HeaderValidationError([
            "No headers found in Excel file. Include columns such as 'Name', 'Set Code', 'Collector Number'."
        ])

    # Yield dict rows lazily
    def _gen():
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = ["" if v is None else str(v).strip() for v in row]
            # Truncate/pad to header length
            if len(vals) < len(headers):
                vals += [""] * (len(headers) - len(vals))
            elif len(vals) > len(headers):
                vals = vals[: len(headers)]
            # Skip fully empty
            if not any(vals):
                continue
            yield {headers[i]: vals[i] for i in range(len(headers))}
    return _gen(), headers

def _open_table(filepath: str) -> tuple[Iterable[Dict[str, Any]], list[str], Optional[str]]:
    """
    Unified row source for CSV or Excel.
    Returns (row_iter, headers, delimiter_if_csv_else_None).
    """
    if _is_excel(filepath):
        rows, headers = _iter_excel_rows(filepath)
        return rows, headers, None
    else:
        reader, delimiter = _make_reader(filepath)
        if not reader.fieldnames:
            raise HeaderValidationError([
                "No headers found in CSV. Include columns such as 'Name', 'Set Code', 'Collector Number'."
            ])
        # Adapt csv.DictReader to our common iterable interface
        def _gen():
            for r in reader:
                yield r
        return _gen(), list(reader.fieldnames), delimiter

# --- PREVIEW API (raw rows with original headers) ---
def preview_csv(filepath: str, default_folder: str = "Unsorted", max_rows: int = 100) -> PreviewData:
    rows_iter, headers, _delimiter = _open_table(filepath)
    if not headers:
        raise HeaderValidationError([
            "No headers found. Include columns such as 'Name', 'Set Code', 'Collector Number'."
        ])

    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        if not any((str(v or "").strip() for v in row.values())):
            continue
        rows.append({h: row.get(h, "") for h in headers})
        if len(rows) >= max_rows:
            break

    return PreviewData(headers=headers, rows=rows)
# --- END PREVIEW API ---


def _log_import_summary(
    stats: ImportStats,
    per_folder: Dict[str, int],
    filepath: str,
    *,
    dry_run: bool,
    quantity_mode: str,
) -> None:
    top = ", ".join(f"{k}:{v}" for k, v in list(per_folder.items())[:5])
    log = current_app.logger if current_app else logger
    log.info(
        "CSV import %s complete: file=%s quantity_mode=%s added=%s updated=%s skipped=%s errors=%s%s",
        "dry-run" if dry_run else "apply",
        filepath,
        quantity_mode,
        stats.added,
        stats.updated,
        stats.skipped,
        stats.errors,
        f"; top folders: {top}" if top else "",
    )
