"""File/header/table parsing helpers for CSV and Excel card imports."""

from __future__ import annotations

import csv
import io
import logging
import os
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from flask import current_app, has_app_context


logger = logging.getLogger(__name__)
IMPORT_MAX_BYTES = int(os.getenv("IMPORT_MAX_BYTES", 10 * 1024 * 1024))  # 10MB default


class HeaderValidationError(ValueError):
    """Raised when required headers are missing from the import file."""

    def __init__(self, details: List[str]):
        super().__init__("Missing required column(s): " + "; ".join(details))
        self.details = details


class FileValidationError(ValueError):
    """Raised when the import file fails size or encoding checks."""


# Ordered header variants (first match wins).
EXPECTED = {
    "folder": [
        "folder name", "folder_name", "folder", "binder", "binder name", "album",
    ],
    "folder_category": [
        "folder category", "folder type", "binder type", "collection type",
    ],
    "name": [
        "card name", "card_name", "name", "card",
    ],
    "qty": [
        "quantity", "qty", "trade quantity", "count", "copies",
    ],
    "set_code": [
        "set code", "set_code", "set", "expansion", "setcode", "edition",
    ],
    "collector_number": [
        "collector number", "collector_number", "collector #",
        "card number", "card_number", "card #",
        "cn", "number", "#",
    ],
    "lang": [
        "language", "lang",
    ],
    "is_foil": [
        "printing", "foil", "is foil", "foil?", "is_foil",
    ],
}

REQUIRED_FIELDS = {
    "name": "Card name",
    "set_code": "Set code",
    "collector_number": "Collector number",
}


def _active_logger():
    return current_app.logger if has_app_context() else logger


def _validate_file_size(filepath: str) -> None:
    try:
        size = Path(filepath).stat().st_size
    except Exception as exc:
        _active_logger().warning("Unable to read import file size for %s: %s", filepath, exc)
        raise FileValidationError("Unable to read import file size.") from exc
    if size > IMPORT_MAX_BYTES:
        _active_logger().warning("Import file too large: %s (%s bytes)", filepath, size)
        raise FileValidationError(
            f"File is too large. Limit: {IMPORT_MAX_BYTES // (1024 * 1024)} MB."
        )


def _read_text(filepath: str) -> str:
    _validate_file_size(filepath)
    data = Path(filepath).read_bytes()
    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        _active_logger().warning("CSV encoding error for %s: %s", filepath, exc)
        raise FileValidationError("CSV must be UTF-8 encoded.") from exc


def _strip_sep_preamble(text: str) -> Tuple[str, Optional[str]]:
    """
    Handle Excel-style `sep=,` preambles, quoted or unquoted.

    Returns ``(content_without_preamble, delimiter_if_declared)``.
    """
    if not text:
        return text, None
    first_newline = text.find("\n")
    head = text if first_newline == -1 else text[:first_newline]
    head_stripped = head.strip().lower().strip("'\"")
    if head_stripped.startswith("sep=") and len(head_stripped) >= 5:
        delimiter = head_stripped[4:5]
        content = text[first_newline + 1:] if first_newline != -1 else ""
        return content, delimiter
    return text, None


def _make_reader(filepath: str) -> Tuple[csv.DictReader, str]:
    raw = _read_text(filepath)
    content, declared = _strip_sep_preamble(raw)

    if declared:
        delimiter = declared
    else:
        sample = content[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
            delimiter = dialect.delimiter
        except Exception:
            delimiter = ","

    sio = io.StringIO(content)
    reader = csv.DictReader(sio, delimiter=delimiter)
    return reader, delimiter


def resolve_header_mapping(
    headers: Iterable[str],
    mapping_override: Optional[Dict[str, str]] = None,
    *,
    allow_missing: bool = False,
) -> Tuple[Dict[str, str], List[str], List[Tuple[str, str]]]:
    if not headers:
        raise HeaderValidationError([
            "No headers found. Include columns such as 'Name', 'Set Code', 'Collector Number'."
        ])
    lower_to_original = {header.strip().lower(): header for header in headers if isinstance(header, str)}
    mapping: Dict[str, str] = {}
    invalid_overrides: List[Tuple[str, str]] = []

    if mapping_override:
        for field, header in mapping_override.items():
            if not header:
                continue
            header_key = str(header).strip()
            if not header_key:
                continue
            resolved = lower_to_original.get(header_key.lower())
            if not resolved:
                invalid_overrides.append((field, header_key))
                continue
            mapping[field] = resolved

    for field, variants in EXPECTED.items():
        if field in mapping:
            continue
        for variant in variants:
            if variant in lower_to_original:
                mapping[field] = lower_to_original[variant]
                break

    missing = [field for field in REQUIRED_FIELDS if field not in mapping]
    if (missing or invalid_overrides) and not allow_missing:
        raise HeaderValidationError(_format_header_errors(missing, invalid_overrides))
    return mapping, missing, invalid_overrides


def _format_header_errors(
    missing: Iterable[str],
    invalid_overrides: Iterable[Tuple[str, str]],
) -> List[str]:
    details: List[str] = []
    for field, header in invalid_overrides:
        label = REQUIRED_FIELDS.get(field, field.replace("_", " ").title())
        details.append(f"{label} mapped to '{header}', but that header is not in the file.")
    for field in missing:
        label = REQUIRED_FIELDS.get(field, field.replace("_", " ").title())
        variants = ", ".join(EXPECTED.get(field, []))
        detail = f"{label} (accepted: {variants})" if variants else label
        details.append(detail)
    return details


def validate_import_file(filepath: str, mapping_override: Optional[Dict[str, str]] = None) -> None:
    """Ensure the provided file has the required headers."""
    _, headers, _ = open_table(filepath)
    resolve_header_mapping(headers or [], mapping_override, allow_missing=False)


def read_import_headers(filepath: str) -> List[str]:
    _, headers, _ = open_table(filepath)
    return headers or []


def is_excel(filepath: str) -> bool:
    return Path(filepath).suffix.lower() in {".xlsx", ".xlsm"}


def _iter_excel_rows(filepath: str) -> tuple[Iterable[Dict[str, Any]], list[str]]:
    """
    Stream rows from the first worksheet of an .xlsx/.xlsm file.

    Returns ``(row_iter, headers)``. All values are coerced to ``str``.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    headers = None
    for row in ws.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if any(values):
            headers = values
            break
    if not headers:
        raise HeaderValidationError([
            "No headers found in Excel file. Include columns such as 'Name', 'Set Code', 'Collector Number'."
        ])

    def _gen():
        for row in ws.iter_rows(min_row=2, values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            if len(values) < len(headers):
                values += [""] * (len(headers) - len(values))
            elif len(values) > len(headers):
                values = values[: len(headers)]
            if not any(values):
                continue
            yield {headers[i]: values[i] for i in range(len(headers))}

    return _gen(), headers


def open_table(filepath: str) -> tuple[Iterable[Dict[str, Any]], list[str], Optional[str]]:
    """
    Unified row source for CSV or Excel.

    Returns ``(row_iter, headers, delimiter_if_csv_else_None)``.
    """
    _validate_file_size(filepath)
    if is_excel(filepath):
        rows, headers = _iter_excel_rows(filepath)
        return rows, headers, None

    reader, delimiter = _make_reader(filepath)
    if not reader.fieldnames:
        raise HeaderValidationError([
            "No headers found in CSV. Include columns such as 'Name', 'Set Code', 'Collector Number'."
        ])

    def _gen():
        for row in reader:
            yield row

    return _gen(), list(reader.fieldnames), delimiter


def count_rows(filepath: str) -> Optional[int]:
    try:
        if is_excel(filepath):
            from openpyxl import load_workbook

            wb = load_workbook(filepath, read_only=True, data_only=True)
            try:
                ws = wb.worksheets[0]
                rows = ws.iter_rows(values_only=True)
                header_found = False
                count = 0
                for row in rows:
                    values = ["" if value is None else str(value).strip() for value in row]
                    if not header_found:
                        if any(values):
                            header_found = True
                        continue
                    if not any(values):
                        continue
                    count += 1
                return count
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
        reader, _delimiter = _make_reader(filepath)
        count = 0
        for row in reader:
            if not any((str(value or "").strip() for value in row.values())):
                continue
            count += 1
        return count
    except Exception:
        return None
